#!/usr/bin/env python3
"""Import the curated Google Sheet ``The Nihali database``.

The workbook is a reviewed integration layer over Nagaraja (2014), Mundlay
(1996), Bhattacharya (1957), and Konow (1906).  It replaces the earlier
Mundlay OCR and Wiktionary-derived Nagaraja installed CSVs while preserving
their entry keys whenever a conservative one-to-one match can be established.

Download the pinned source snapshot with::

    curl -L 'https://docs.google.com/spreadsheets/d/\
1Mas3uqXcpXAFPV__OMv_uGPXb_d67iaMhqAQGITisrg/export?format=xlsx' \
      -o /tmp/nihali-database.xlsx

Then preview and install from the ``data/`` directory with::

    uv run --with openpyxl python data/other/forms/raw_data/nihali_database.py \
      --xlsx /tmp/nihali-database.xlsx
    uv run --with openpyxl python data/other/forms/raw_data/nihali_database.py \
      --xlsx /tmp/nihali-database.xlsx --install
"""

from __future__ import annotations

import argparse
import csv
import difflib
import hashlib
import json
import re
import tempfile
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
DATA_ROOT = HERE.parents[3]
FORMS_DIR = DATA_ROOT / "data/other/forms"
SNAPSHOT_SHA256 = "a2525d858969c84eb36c4f5a43857a893b89baa1e0bee16974bc4e8a9d46524d"
SNAPSHOT_EXPORTED = "2026-08-17"
DRIVE_MODIFIED = "2026-04-22T02:38:24.316Z"
SPREADSHEET_ID = "1Mas3uqXcpXAFPV__OMv_uGPXb_d67iaMhqAQGITisrg"
EXPECTED_ACTIVE = {
    "Nagaraja": 1698,
    "Mundlay": 1706,
    "Bhattacharya": 384,
    "Konow": 190,
    "Contact": 34,
    "Roots": 1,
    "Dravidian": 22,
}
OUTPUT_NAMES = {
    "Nagaraja": "20260817-nagaraja-nihali-wiktionary.csv",
    "Mundlay": "20260817-mundlay-nihali.csv",
    "Bhattacharya": "20260817-nihali-database-bhattacharya.csv",
    "Konow": "20260817-nihali-database-konow.csv",
}
AUDIT_NAME = "20260817-nihali-database-audit.csv"
KEY_MAP_NAME = "20260817-nihali-database-key-map.csv"


def compact(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", str(value))).strip()


def source_id(value: object) -> str:
    value = compact(value)
    if re.fullmatch(r"\d+\.0", value):
        return value[:-2]
    return value


def fold(value: str) -> str:
    value = unicodedata.normalize("NFKD", compact(value)).casefold()
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.translate(
        str.maketrans(
            {
                "ʈ": "t", "ṭ": "t", "ɖ": "d", "ḍ": "d", "ɽ": "r", "ṛ": "r",
                "ɳ": "n", "ṇ": "n", "ŋ": "n", "ñ": "n", "ʃ": "s", "ś": "s",
                "č": "c", "ʔ": "", "ː": "", "w": "v", "ᵑ": "n",
            }
        )
    )
    return re.sub(r"[^a-z0-9]+", "", value)


def fold_gloss(value: str) -> str:
    value = unicodedata.normalize("NFKD", compact(value)).casefold()
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def split_variants(value: str) -> list[tuple[str, str]]:
    """Return source-form variants plus any detached transcription annotation."""
    parts, current, depth = [], [], 0
    for char in compact(value):
        if char == "(":
            depth += 1
        elif char == ")" and depth:
            depth -= 1
        if char in ",;" and depth == 0:
            parts.append("".join(current))
            current = []
        else:
            current.append(char)
    parts.append("".join(current))
    result = []
    expanded_parts = []
    for part in parts:
        expanded_parts.extend(re.split(r"\s+~\s+", part))
    for part in expanded_parts:
        part = compact(part).strip(" ,;")
        if not part:
            continue
        annotation = ""
        match = re.search(r"\s+(\([^)]{1,24}\))$", part)
        if match:
            annotation = match.group(1)
            part = part[: match.start()].strip()
        part = part.strip(" ‘ʻ’\"“”")
        if part:
            result.append((part, annotation))
    return result


def active_rows(sheet) -> list[tuple[int, tuple[object, ...]]]:
    rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if any(value not in (None, "") for value in row):
            rows.append((row_number, tuple(row)))
    return rows


@dataclass
class Legacy:
    key: str
    form: str
    gloss: str
    page: str


def read_legacy(path: Path) -> list[Legacy]:
    if not path.exists():
        return []
    entries = []
    with path.open(encoding="utf-8", newline="") as stream:
        for row in csv.reader(stream):
            if len(row) < 11:
                continue
            page_match = re.search(r"p\.\s*(\d+)", row[7])
            entries.append(Legacy(row[10], row[2], row[3], page_match.group(1) if page_match else ""))
    return entries


def match_legacy(
    records: list[tuple[str, int, int, str, str, str]],
    legacy: list[Legacy],
) -> dict[tuple[str, int, int], tuple[str, float, str]]:
    """Conservatively reconcile (tab,row,variant,form,gloss,page) to old keys."""
    used = set()
    matches = {}
    for tab, row_number, variant, form, gloss, page in records:
        scored = []
        ff, fg = fold(form), fold_gloss(gloss)
        for index, old in enumerate(legacy):
            if index in used or (page and old.page and page != old.page):
                continue
            of, og = fold(old.form), fold_gloss(old.gloss)
            form_score = 1.0 if ff and ff == of else difflib.SequenceMatcher(None, ff, of).ratio()
            gloss_score = 1.0 if fg and fg == og else difflib.SequenceMatcher(None, fg, og).ratio()
            score = 0.72 * form_score + 0.28 * gloss_score
            scored.append((score, form_score, gloss_score, index, old))
        if not scored:
            continue
        scored.sort(key=lambda item: (-item[0], -item[1], -item[2], item[4].key))
        best = scored[0]
        margin = best[0] - scored[1][0] if len(scored) > 1 else 1.0
        exact = best[1] == 1.0 and best[2] == 1.0
        accepted = exact or (
            best[0] >= 0.88 and best[1] >= 0.82 and margin >= 0.025
        )
        if accepted:
            used.add(best[3])
            reason = "exact_form_gloss" if exact else "conservative_fuzzy"
            matches[(tab, row_number, variant)] = (best[4].key, best[0], reason)
    return matches


def load_key_map(path: Path) -> dict[tuple[str, int, int], tuple[str, float, str]]:
    if not path.exists():
        return {}
    result = {}
    with path.open(encoding="utf-8", newline="") as stream:
        for row in csv.DictReader(stream):
            if row["Legacy_Key"]:
                result[(row["Tab"], int(row["Sheet_Row"]), int(row["Variant"]))] = (
                    row["Legacy_Key"], float(row["Score"]), row["Reason"]
                )
    return result


def write_key_map(
    path: Path,
    records: list[tuple[str, int, int, str, str, str]],
    matches: dict[tuple[str, int, int], tuple[str, float, str]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=["Tab", "Sheet_Row", "Variant", "Form", "Gloss", "Page", "Legacy_Key", "Score", "Reason"],
        )
        writer.writeheader()
        for tab, row_number, variant, form, gloss, page in records:
            matched = matches.get((tab, row_number, variant), ("", 0.0, "unmatched"))
            writer.writerow(
                {
                    "Tab": tab, "Sheet_Row": row_number, "Variant": variant,
                    "Form": form, "Gloss": gloss, "Page": page,
                    "Legacy_Key": matched[0], "Score": f"{matched[1]:.4f}" if matched[0] else "",
                    "Reason": matched[2],
                }
            )


def exact_etymon(text: str, cdial_ids: set[str], dedr_ids: set[str]) -> tuple[str, str]:
    """Return a conservative marked-borrowing Parameter_ID and audit reason."""
    if re.search(r"(?i)\b(?:possibly|perhaps|maybe|cf\.|uncertain)\b|\?", text):
        return "", "uncertain_etymology"
    cdial = set(re.findall(r"(?i)\bCDIAL\s*([0-9]+[a-z]?)\b", text))
    dedr = {f"d{x}" for x in re.findall(r"(?i)\bDED(?:R)?\s*([0-9]+)\b", text)}
    candidates = ({x for x in cdial if x in cdial_ids} | {x for x in dedr if x in dedr_ids})
    if len(candidates) == 1:
        return ">" + next(iter(candidates)), "exact_printed_id_marked_borrowing"
    if cdial or dedr:
        return "", "ambiguous_or_missing_etymon_id"
    return "", "no_resolvable_etymon_id"


def combine_labeled(parts: Iterable[tuple[str, str]]) -> str:
    return "; ".join(f"{label}: {compact(value)}" for label, value in parts if compact(value))


def dedupe_parts(*values: str) -> str:
    return "; ".join(dict.fromkeys(compact(value) for value in values if compact(value)))


def database_locator(tab: str, row_number: int, sid: str) -> str:
    suffix = f", ID {sid}" if sid else ""
    return f"nihali-database2026[tab {tab}, row {row_number}{suffix}]"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--install", action="store_true")
    parser.add_argument("--allow-new-snapshot", action="store_true")
    parser.add_argument("--preview-dir", type=Path)
    args = parser.parse_args()

    digest = hashlib.sha256(args.xlsx.read_bytes()).hexdigest()
    if digest != SNAPSHOT_SHA256 and not args.allow_new_snapshot:
        raise ValueError(f"snapshot SHA-256 changed: {digest}; expected {SNAPSHOT_SHA256}")
    workbook = load_workbook(args.xlsx, read_only=False, data_only=False)
    if set(workbook.sheetnames) != set(EXPECTED_ACTIVE):
        raise ValueError(f"unexpected tabs: {workbook.sheetnames}")
    tabs = {name: active_rows(workbook[name]) for name in workbook.sheetnames}
    actual = {name: len(rows) for name, rows in tabs.items()}
    if actual != EXPECTED_ACTIVE:
        raise ValueError(f"active-row counts changed: {actual}")

    if args.install:
        output_dir = FORMS_DIR
        audit_path = HERE / AUDIT_NAME
        key_map_path = HERE / KEY_MAP_NAME
    else:
        output_dir = args.preview_dir or Path(tempfile.mkdtemp(prefix="nihali-database-preview-"))
        audit_path = output_dir / AUDIT_NAME
        key_map_path = output_dir / KEY_MAP_NAME
    output_dir.mkdir(parents=True, exist_ok=True)

    # Record each prospective variant before writing, so legacy keys can be
    # matched globally one-to-one rather than greedily within individual rows.
    prospective = []
    for tab in ("Nagaraja", "Mundlay"):
        for row_number, row in tabs[tab]:
            form = compact(row[2] if tab == "Nagaraja" else row[1])
            gloss = compact(row[1] if tab == "Nagaraja" else row[2])
            page = ""
            if tab == "Mundlay":
                match = re.search(r"(\d+)\s*$", compact(row[4]))
                page = match.group(1) if match else ""
            for variant, (variant_form, _) in enumerate(split_variants(form), 1):
                prospective.append((tab, row_number, variant, variant_form, gloss, page))

    canonical_key_map = HERE / KEY_MAP_NAME
    matches = load_key_map(canonical_key_map)
    if not matches:
        matches.update(
            match_legacy(
                [record for record in prospective if record[0] == "Mundlay"],
                read_legacy(FORMS_DIR / OUTPUT_NAMES["Mundlay"]),
            )
        )
        matches.update(
            match_legacy(
                [record for record in prospective if record[0] == "Nagaraja"],
                read_legacy(FORMS_DIR / OUTPUT_NAMES["Nagaraja"]),
            )
        )
    write_key_map(key_map_path, prospective, matches)

    with (DATA_ROOT / "data/cdial/params.csv").open(encoding="utf-8", newline="") as stream:
        cdial_ids = {row[0] for row in csv.reader(stream)}
    with (DATA_ROOT / "data/dedr/params.csv").open(encoding="utf-8", newline="") as stream:
        dedr_ids = {row[0] for row in csv.reader(stream)}

    dravidian_by_id = defaultdict(list)
    for row_number, row in tabs["Dravidian"]:
        sid = source_id(row[0])
        if sid and compact(row[2]):
            dravidian_by_id[sid].append((row_number, row))

    output_streams = {}
    output_writers = {}
    for tab, filename in OUTPUT_NAMES.items():
        stream = (output_dir / filename).open("w", encoding="utf-8", newline="")
        output_streams[tab] = stream
        output_writers[tab] = csv.writer(stream)

    audit_fields = [
        "Status", "Reason", "Tab", "Sheet_Row", "Source_ID", "Raw_Cells_JSON",
        "Parsed_Form", "Parsed_Gloss", "Output_Keys", "Output_Count", "Reference",
        "Parameter_ID", "Etymology_Match_Status", "Legacy_Keys", "Legacy_Match_Scores",
        "Language_ID", "Snapshot_SHA256", "Snapshot_Exported", "Drive_Modified",
    ]
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_stream = audit_path.open("w", encoding="utf-8", newline="")
    audit = csv.DictWriter(audit_stream, fieldnames=audit_fields)
    audit.writeheader()
    counts = Counter()

    try:
        for tab in workbook.sheetnames:
            for row_number, row in tabs[tab]:
                sid = source_id(row[0])
                raw_json = json.dumps([compact(value) for value in row], ensure_ascii=False)
                if tab not in OUTPUT_NAMES:
                    reason = "analysis_sidecar_merged_by_source_id" if tab == "Dravidian" and sid and compact(row[2]) else "nonlexical_analysis_sidecar"
                    audit.writerow(
                        {
                            "Status": "excluded", "Reason": reason, "Tab": tab,
                            "Sheet_Row": row_number, "Source_ID": sid, "Raw_Cells_JSON": raw_json,
                            "Language_ID": "Ni", "Snapshot_SHA256": digest,
                            "Snapshot_Exported": SNAPSHOT_EXPORTED, "Drive_Modified": DRIVE_MODIFIED,
                        }
                    )
                    counts[f"{tab}:excluded"] += 1
                    continue

                if tab == "Nagaraja":
                    gloss, form = compact(row[1]), compact(row[2])
                    printed_etymology, editor_etymology = compact(row[3]), compact(row[4])
                    comments, correspondences = compact(row[5]), compact(row[6])
                    for _dr_row, dr in dravidian_by_id.get(sid, []):
                        if fold(compact(dr[2])) == fold(form):
                            editor_etymology = dedupe_parts(editor_etymology, compact(dr[4]))
                            correspondences = dedupe_parts(correspondences, compact(dr[8]))
                    original_source = "nagaraja2014[pp. 250–332]"
                    etymology = combine_labeled(
                        (("Nagaraja", printed_etymology), ("Database editor", editor_etymology))
                    )
                    notes = combine_labeled((("Comment", comments), ("Sound correspondence", correspondences)))
                elif tab == "Mundlay":
                    form, gloss = compact(row[1]), compact(row[2])
                    printed_etymology, editor_etymology = compact(row[3]), compact(row[9])
                    match = re.search(r"(\d+)\s*$", compact(row[4]))
                    page = match.group(1) if match else ""
                    original_source = f"mundlay1996[p. {page}]" if page else "mundlay1996[pp. 17–40]"
                    etymology = combine_labeled(
                        (("Mundlay", printed_etymology), ("Database editor", editor_etymology))
                    )
                    notes = ""
                elif tab == "Bhattacharya":
                    form, gloss = compact(row[1]), compact(row[2])
                    printed_etymology, editor_etymology = compact(row[3]), ""
                    match = re.search(r"(\d+)\s*$", compact(row[4]))
                    page = match.group(1) if match else ""
                    original_source = f"bhattacharya1957[p. {page}]" if page else "bhattacharya1957[pp. 245–258]"
                    etymology = combine_labeled((("Bhattacharya", printed_etymology),))
                    notes = ""
                else:  # Konow
                    gloss, form = compact(row[1]), compact(row[2])
                    printed_etymology, editor_etymology = compact(row[4]), ""
                    original_source = "konow1906[pp. 185–189]"
                    etymology = combine_labeled((("Database comment", printed_etymology),))
                    notes = ""

                variants = split_variants(form)
                reference = original_source + ";" + database_locator(tab, row_number, sid)
                if not variants:
                    audit.writerow(
                        {
                            "Status": "excluded", "Reason": "blank_or_illegible_form", "Tab": tab,
                            "Sheet_Row": row_number, "Source_ID": sid, "Raw_Cells_JSON": raw_json,
                            "Parsed_Gloss": gloss, "Reference": reference, "Language_ID": "Ni",
                            "Snapshot_SHA256": digest, "Snapshot_Exported": SNAPSHOT_EXPORTED,
                            "Drive_Modified": DRIVE_MODIFIED,
                        }
                    )
                    counts[f"{tab}:excluded"] += 1
                    continue

                param, match_status = exact_etymon(etymology, cdial_ids, dedr_ids)
                uncertain_glyph = any(unicodedata.category(char) == "Co" for char in form)
                uncertain = uncertain_glyph or "?" in etymology or bool(re.search(r"(?i)\b(?:possibly|perhaps|maybe|uncertain)\b", etymology))
                loanword = bool(printed_etymology or editor_etymology)
                tags = " ".join(tag for tag, active in (("loanword", loanword), ("uncertain", uncertain)) if active)
                output_keys, legacy_keys, legacy_scores = [], [], []
                main_key = ""
                for variant, (variant_form, annotation) in enumerate(variants, 1):
                    matched = matches.get((tab, row_number, variant))
                    base = f"nihali-database2026:{tab.casefold()}:{sid or f'r{row_number}'}"
                    key = matched[0] if matched else base if variant == 1 else f"{base}:variant:{variant}"
                    if variant > 1 and key == base:
                        key = f"{base}:variant:{variant}"
                    if not main_key:
                        main_key = key
                    variant_notes = dedupe_parts(notes, f"Source transcription annotation: {annotation}" if annotation else "")
                    output_writers[tab].writerow(
                        [
                            "Ni", param, variant_form, gloss, "", variant_form, variant_notes,
                            reference, "", etymology, key, main_key if variant > 1 else "", "", "", tags,
                        ]
                    )
                    output_keys.append(key)
                    if matched:
                        legacy_keys.append(matched[0])
                        legacy_scores.append(f"{matched[1]:.4f}")
                    counts[f"{tab}:installed"] += 1
                audit.writerow(
                    {
                        "Status": "ingested", "Reason": (
                            "curated_spreadsheet_record;private_use_glyph_preserved"
                            if uncertain_glyph else "curated_spreadsheet_record"
                        ), "Tab": tab,
                        "Sheet_Row": row_number, "Source_ID": sid, "Raw_Cells_JSON": raw_json,
                        "Parsed_Form": form, "Parsed_Gloss": gloss, "Output_Keys": "|".join(output_keys),
                        "Output_Count": len(output_keys), "Reference": reference, "Parameter_ID": param,
                        "Etymology_Match_Status": match_status, "Legacy_Keys": "|".join(legacy_keys),
                        "Legacy_Match_Scores": "|".join(legacy_scores), "Language_ID": "Ni",
                        "Snapshot_SHA256": digest, "Snapshot_Exported": SNAPSHOT_EXPORTED,
                        "Drive_Modified": DRIVE_MODIFIED,
                    }
                )
                counts[f"{tab}:source"] += 1
    finally:
        audit_stream.close()
        for stream in output_streams.values():
            stream.close()

    print(f"snapshot: {digest}")
    print(f"output directory: {output_dir}")
    for key in sorted(counts):
        print(f"{key}: {counts[key]}")
    print(f"legacy keys retained: {len(matches)}")
    print(f"audit: {audit_path}")
    print(f"key map: {key_map_path}")


if __name__ == "__main__":
    main()
